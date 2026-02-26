"""
O'quvchilar uchun utility funksiyalari.
"""
from openpyxl import load_workbook
from datetime import datetime
from typing import List, Dict, Any, Optional
import logging

logger = logging.getLogger(__name__)


def parse_excel_file(file) -> List[Dict[str, Any]]:
    """
    Excel faylni o'qib, o'quvchilar ro'yxatini qaytaradi.
    
    Fayl strukturasi:
    0: Shartnoma raqami (ishlatilmaydi)
    1: FIO (Familiya Ism Otasining_ismi)
    2: Balans
    3: Sinf
    4: Guruh (ishlatilmaydi)
    5: Telefon raqam
    6: Sinf rahbar (ishlatilmaydi)
    7: Jinsi
    8: Tug'ilgan kuni
    9: Manzil
    
    Args:
        file: Django UploadedFile obyekti yoki fayl yo'li (string)
        
    Returns:
        List[Dict]: O'quvchilar ma'lumotlari ro'yxati
    """
    try:
        # File path string bo'lsa, o'qish
        if isinstance(file, str):
            workbook = load_workbook(file, data_only=True)
        else:
            # File object bo'lsa, to'g'ridan-to'g'ri o'qish
            workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
        
        students_data = []
        
        # MUHIM: Birinchi qatorni (header) o'qish va logging qilish
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        logger.info(f"Excel header columns: {[str(col) for col in header_row]}")
        
        # Headerdan keyin ma'lumotlarni o'qiymiz (2-qatordan boshlab)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):  # Bo'sh qatorni o'tkazib yuborish
                continue
            
            # DEBUGGING: Birinchi qatorni to'liq ko'rsatish
            if row_idx == 2:
                logger.info(f"Birinchi student qatori (row {row_idx}): {[str(cell) if cell is not None else 'NULL' for cell in row[:10]]}")
                
            # Column 1: FIO - Familiya, Ism, Otasining ismi
            full_name = str(row[1] or '').strip()
            if not full_name or full_name in ('None', 'NULL', 'null', ''):
                logger.warning(f"Qator {row_idx}: FIO bo'sh, o'tkazib yuborilmoqda")
                continue
            
            name_parts = full_name.split()
            if len(name_parts) < 2:
                logger.warning(f"Qator {row_idx}: FIO kamida 2 qismdan iborat bo'lishi kerak: '{full_name}'")
                continue
                
            # Uzbek formatda: Familiya Ism Otasining_ismi
            last_name = name_parts[0]  # Familiya
            first_name = name_parts[1] if len(name_parts) > 1 else ''  # Ism
            middle_name = name_parts[2] if len(name_parts) > 2 else ''  # Otasining ismi
            
            # Column 5: Telefon raqam
            phone_raw = row[5]
            if phone_raw is None or str(phone_raw).strip() in ('None', 'NULL', '', 'null'):
                logger.warning(f"Qator {row_idx}: Telefon bo'sh, o'tkazib yuborilmoqda")
                continue
            
            # Telefon raqamni tozalash va formatlash
            phone = _clean_phone(str(phone_raw))
            if not phone:
                logger.warning(f"Qator {row_idx}: Telefon raqam noto'g'ri: '{phone_raw}'")
                continue
            
            # Column 2: Balans
            balance = _parse_balance(row[2])
            
            # Column 3: Sinf nomi
            class_name = str(row[3] or '').strip()
            if not class_name or class_name in ('None', 'NULL', 'null', ''):
                class_name = None
            
            # Column 7: Jinsi
            gender_value = str(row[7] or 'male').strip().lower()
            gender = 'male' if gender_value in ['male', 'erkak', 'm', 'e'] else 'female' if gender_value in ['female', 'ayol', 'f', 'a'] else 'unspecified'
            
            # Column 8: Tug'ilgan sana
            birth_date = None
            birth_date_value = row[8]
            if birth_date_value:
                try:
                    if isinstance(birth_date_value, datetime):
                        birth_date = birth_date_value.date()
                    elif isinstance(birth_date_value, str):
                        # Turli formatlarni qo'llab-quvvatlash
                        for fmt in ['%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y/%m/%d']:
                            try:
                                birth_date = datetime.strptime(birth_date_value.strip(), fmt).date()
                                break
                            except ValueError:
                                continue
                except Exception as e:
                    logger.warning(f"Qator {row_idx}: Tug'ilgan sanani parse qilishda xatolik: {e}")
            
            # Column 9: Manzil
            address = str(row[9] or '').strip()
            
            student_data = {
                'row_number': row_idx,
                'first_name': first_name,
                'middle_name': middle_name,
                'last_name': last_name,
                'phone_number': phone,
                'gender': gender,
                'date_of_birth': birth_date,
                'address': address,
                'class_name': class_name,
                'balance': balance,
            }
            
            students_data.append(student_data)
        
        logger.info(f"Excel fayldan {len(students_data)} ta o'quvchi ma'lumoti o'qildi")
        return students_data
        
    except Exception as e:
        logger.error(f"Excel faylni o'qishda xatolik: {str(e)}")
        raise ValueError(f"Excel faylni o'qishda xatolik: {str(e)}")


def _parse_balance(value) -> float:
    """Balansni parse qilish."""
    if not value:
        return 0.0
    try:
        # String bo'lsa, bo'sh joylarni olib tashlaymiz
        if isinstance(value, str):
            value = value.replace(' ', '').replace(',', '.')
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _clean_phone(phone: str) -> str:
    """
    Telefon raqamni tozalash va formatlash.
    
    Excel faylda telefon raqamlar turli formatlarda bo'lishi mumkin:
    - 999971000081 (12 raqam, birinchi 9 xato)
    - 99971000081 (11 raqam, 999 o'rniga 998 bo'lishi kerak)
    - 998901234567 (to'g'ri format)
    - +998901234567 (to'g'ri format)
    """
    if not phone:
        return ''
    
    # Bo'sh joylar va maxsus belgilarni olib tashlash
    phone = str(phone).strip().replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('+', '')
    
    # Faqat raqamlarni qoldirish
    phone = ''.join(c for c in phone if c.isdigit())
    
    # Agar telefon raqam 12 raqamli bo'lib, 9999... bilan boshlansa, 
    # birinchi 9 ni olib tashlash (xato kiritilgan bo'lishi mumkin)
    if len(phone) == 12 and phone.startswith('9999'):
        phone = phone[1:]  # 999971000081 -> 99971000081
        logger.info(f"12 raqamli telefon tuzatildi: {phone}")
    
    # Agar 11 raqamli bo'lib, 999... bilan boshlansa, 998 ga o'zgartirish
    if len(phone) == 11 and phone.startswith('999'):
        phone = '998' + phone[3:]  # 99971000081 -> 99871000081
        logger.info(f"11 raqamli telefon tuzatildi: {phone}")
    
    # Agar 9 raqamli bo'lib, 9... bilan boshlansa (masalan 901234567), 998 qo'shamiz
    if len(phone) == 9 and phone.startswith('9'):
        phone = '998' + phone
    
    # + qo'shish
    if phone and not phone.startswith('+'):
        phone = '+' + phone
    
    return phone


def _map_relationship_type(rel_type: str) -> str:
    """
    Yaqinlik turini mapping qilish.
    
    Excel faylda turli nomlar bo'lishi mumkin, ularni standart turlarga o'tkazamiz.
    """
    if not rel_type:
        return 'guardian'
    
    rel_type = rel_type.lower().strip()
    
    # Mapping dictionary
    mapping = {
        'ota': 'father',
        'otasi': 'father',
        'dada': 'father',
        'father': 'father',
        
        'ona': 'mother',
        'onasi': 'mother',
        'oyi': 'mother',
        'mother': 'mother',
        
        'aka': 'brother',
        'ukasi': 'brother',
        'akasi': 'brother',
        'brother': 'brother',
        
        'opa': 'sister',
        'singil': 'sister',
        'opasi': 'sister',
        'sister': 'sister',
        
        'bobo': 'grandfather',
        'bobosi': 'grandfather',
        'grandfather': 'grandfather',
        
        'buvi': 'grandmother',
        'buvisi': 'grandmother',
        'grandmother': 'grandmother',
        
        'amaki': 'uncle',
        'tog\'a': 'uncle',
        'uncle': 'uncle',
        
        'xola': 'aunt',
        'amma': 'aunt',
        'aunt': 'aunt',
        
        'vasiy': 'guardian',
        'guardian': 'guardian',
    }
    
    return mapping.get(rel_type, 'guardian')


def create_student_balance_transaction(branch, student_profile, amount: int):
    """
    Student uchun boshlang'ich balans tranzaksiyasini yaratish.
    
    Bu funksiya:
    1. Transaction yaratadi (INCOME type, COMPLETED status)
    2. StudentBalance.add_amount() atomic metodini ishlatadi
    3. Race condition xavfidan himoyalangan
    
    Args:
        branch: Branch instance
        student_profile: StudentProfile instance
        amount: Balans miqdori (so'm, musbat integer)
        
    Returns:
        Transaction instance
        
    Raises:
        ValueError: amount manfiy yoki 0 bo'lsa
        Exception: StudentBalance topilmasa yoki boshqa xatolikda
    """
    from apps.school.finance.models import (
        Transaction, TransactionType, TransactionStatus,
        PaymentMethod, CashRegister, StudentBalance
    )
    from django.db import transaction as db_transaction
    from django.utils import timezone
    
    if amount <= 0:
        raise ValueError(f"Amount musbat bo'lishi kerak: {amount}")
    
    # StudentBalance mavjudligini tekshirish
    try:
        student_balance = StudentBalance.objects.get(student_profile=student_profile)
    except StudentBalance.DoesNotExist:
        raise Exception(
            f"StudentBalance topilmadi: student_profile_id={student_profile.id}, "
            f"phone={student_profile.user_branch.user.phone_number}"
        )
    
    # Asosiy kassani topish yoki yaratish
    cash_register, _ = CashRegister.objects.get_or_create(
        branch=branch,
        name="Asosiy kassa",
        defaults={
            'description': 'Avtomatik yaratilgan asosiy kassa',
            'is_active': True,
            'balance': 0
        }
    )
    
    # Atomic transaction ichida tranzaksiya yaratish va balansni yangilash
    with db_transaction.atomic():
        # 1. Transaction yaratish
        txn = Transaction.objects.create(
            branch=branch,
            cash_register=cash_register,
            transaction_type=TransactionType.INCOME,
            status=TransactionStatus.COMPLETED,
            amount=amount,
            payment_method=PaymentMethod.CASH,
            description=f"Import - boshlang'ich balans: {student_profile}",
            student_profile=student_profile,
            transaction_date=timezone.now(),
            metadata={
                'source': 'student_import',
                'import_date': timezone.now().isoformat(),
                'student_phone': student_profile.user_branch.user.phone_number,
                'personal_number': student_profile.personal_number
            }
        )
        
        # 2. StudentBalance yangilash (atomic operation)
        student_balance.add_amount(amount)
    
    logger.info(
        f"Balance tranzaksiya yaratildi: Student={student_profile}, "
        f"Amount={amount}, Transaction ID={txn.id}"
    )
    
    return txn
