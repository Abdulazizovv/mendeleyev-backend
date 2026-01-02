"""
Moliya tizimi uchun Celery tasklar.
"""
import os
from datetime import datetime
from io import BytesIO
import logging
from celery import shared_task
from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


@shared_task(bind=True, name='finance.export_transactions_to_excel')
def export_transactions_to_excel(self, branch_id, filters=None, user_id=None):
    """
    Tranzaksiyalarni Excel fayliga export qilish.
    
    Args:
        branch_id: Filial ID
        filters: Filter parametrlari (dict)
        user_id: Export qiluvchi user ID
        
    Returns:
        dict: {
            'success': bool,
            'file_path': str,  # Media folder ichidagi fayl yo'li
            'file_url': str,   # To'liq URL
            'filename': str,
            'records_count': int,
            'error': str (agar xatolik bo'lsa)
        }
    """
    try:
        from apps.school.finance.models import Transaction, TransactionStatus
        from apps.branch.models import Branch
        from django.conf import settings
        
        # Branch ma'lumotini olish
        try:
            branch = Branch.objects.get(id=branch_id, deleted_at__isnull=True)
        except Branch.DoesNotExist:
            return {
                'success': False,
                'error': 'Filial topilmadi'
            }
        
        # Tranzaksiyalarni olish
        queryset = Transaction.objects.filter(
            branch_id=branch_id,
            deleted_at__isnull=True
        ).select_related(
            'branch',
            'cash_register',
            'category',
            'student_profile',
            'student_profile__user_branch__user',
            'employee_membership',
            'employee_membership__user'
        ).order_by('-transaction_date', '-created_at')
        
        logger.info(f"Export task started: branch_id={branch_id}, filters={filters}")
        logger.info(f"Base queryset count (before filters): {queryset.count()}")
        
        # Filterlarni qo'llash
        if filters:
            # Transaction type filter
            if 'transaction_type' in filters:
                queryset = queryset.filter(transaction_type=filters['transaction_type'])
            
            # Status filter
            if 'status' in filters:
                queryset = queryset.filter(status=filters['status'])
            
            # Date range filter
            if 'date_from' in filters:
                queryset = queryset.filter(transaction_date__gte=filters['date_from'])
            if 'date_to' in filters:
                queryset = queryset.filter(transaction_date__lte=filters['date_to'])
            
            # Cash register filter
            if 'cash_register' in filters:
                queryset = queryset.filter(cash_register_id=filters['cash_register'])
            
            # Category filter
            if 'category' in filters:
                queryset = queryset.filter(category_id=filters['category'])
            
            # Student filter
            if 'student_profile' in filters:
                queryset = queryset.filter(student_profile_id=filters['student_profile'])
        
        # Limitatsiya: maksimum 50,000 yozuv
        MAX_RECORDS = 50000
        records_count = queryset.count()
        
        logger.info(f"Final queryset count (after filters): {records_count}")
        
        if records_count > MAX_RECORDS:
            logger.warning(f"Too many records for export: {records_count}")
            return {
                'success': False,
                'error': f'Juda ko\'p yozuv ({records_count}). Maksimum {MAX_RECORDS} yozuvgacha export qilish mumkin. Iltimos, sana oralig\'ini qisqartiring.'
            }
        
        if records_count == 0:
            logger.warning(f"No records found for export. Branch: {branch_id}, Filters: {filters}")
            return {
                'success': False,
                'error': 'Export qilish uchun ma\'lumot topilmadi'
            }
        
        # Excel fayl yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = "Tranzaksiyalar"
        
        # Headerlar
        headers = [
            '№',
            'Sana',
            'Turi',
            'Holat',
            'Summa (so\'m)',
            'To\'lov usuli',
            'Kassa',
            'Kategoriya',
            'O\'quvchi',
            'Xodim',
            'Tavsif',
            'Referens raqam',
            'Yaratilgan',
        ]
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headerlarni yozish
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Ma'lumotlarni yozish
        for row_num, transaction in enumerate(queryset, 2):
            # 1. Tartib raqam
            ws.cell(row=row_num, column=1, value=row_num - 1).border = border
            
            # 2. Sana
            date_value = transaction.transaction_date.strftime('%d.%m.%Y %H:%M') if transaction.transaction_date else ''
            ws.cell(row=row_num, column=2, value=date_value).border = border
            
            # 3. Turi
            ws.cell(row=row_num, column=3, value=transaction.get_transaction_type_display()).border = border
            
            # 4. Holat
            status_cell = ws.cell(row=row_num, column=4, value=transaction.get_status_display())
            status_cell.border = border
            # Status rangini belgilash
            if transaction.status == TransactionStatus.COMPLETED:
                status_cell.font = Font(color="008000")  # Yashil
            elif transaction.status == TransactionStatus.PENDING:
                status_cell.font = Font(color="FFA500")  # Orange
            elif transaction.status == TransactionStatus.FAILED:
                status_cell.font = Font(color="FF0000")  # Qizil
            
            # 5. Summa
            amount_cell = ws.cell(row=row_num, column=5, value=transaction.amount)
            amount_cell.number_format = '#,##0'
            amount_cell.border = border
            
            # 6. To'lov usuli
            payment_method = transaction.get_payment_method_display() if transaction.payment_method else ''
            ws.cell(row=row_num, column=6, value=payment_method).border = border
            
            # 7. Kassa
            cash_register_name = transaction.cash_register.name if transaction.cash_register else ''
            ws.cell(row=row_num, column=7, value=cash_register_name).border = border
            
            # 8. Kategoriya
            category_name = transaction.category.name if transaction.category else ''
            ws.cell(row=row_num, column=8, value=category_name).border = border
            
            # 9. O'quvchi
            student_name = ''
            if transaction.student_profile and hasattr(transaction.student_profile, 'user_branch'):
                user = transaction.student_profile.user_branch.user
                student_name = f"{user.first_name} {user.last_name}".strip()
                personal_number = transaction.student_profile.personal_number
                if personal_number:
                    student_name = f"{student_name} ({personal_number})"
            ws.cell(row=row_num, column=9, value=student_name).border = border
            
            # 10. Xodim
            employee_name = ''
            if transaction.employee_membership:
                user = transaction.employee_membership.user
                employee_name = f"{user.first_name} {user.last_name}".strip()
            ws.cell(row=row_num, column=10, value=employee_name).border = border
            
            # 11. Tavsif
            description = transaction.description or ''
            ws.cell(row=row_num, column=11, value=description).border = border
            
            # 12. Referens raqam
            reference = transaction.reference_number or ''
            ws.cell(row=row_num, column=12, value=reference).border = border
            
            # 13. Yaratilgan
            created_date = transaction.created_at.strftime('%d.%m.%Y %H:%M') if transaction.created_at else ''
            ws.cell(row=row_num, column=13, value=created_date).border = border
        
        # Ustunlar kengligini avtomatik moslash
        column_widths = {
            1: 8,   # №
            2: 18,  # Sana
            3: 15,  # Turi
            4: 12,  # Holat
            5: 15,  # Summa
            6: 15,  # To'lov usuli
            7: 20,  # Kassa
            8: 20,  # Kategoriya
            9: 30,  # O'quvchi
            10: 25, # Xodim
            11: 35, # Tavsif
            12: 20, # Referens
            13: 18, # Yaratilgan
        }
        
        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Faylni saqlash
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"transactions_{branch.name}_{timestamp}.xlsx"
        
        # Media/exports papkasini yaratish
        exports_dir = os.path.join(settings.MEDIA_ROOT, 'exports', 'finance')
        os.makedirs(exports_dir, exist_ok=True)
        
        file_path = os.path.join(exports_dir, filename)
        
        # Excel faylni saqlash
        wb.save(file_path)
        
        # Relative path (media root dan keyin)
        relative_path = os.path.join('exports', 'finance', filename)
        file_url = f"{settings.MEDIA_URL}{relative_path}"
        
        logger.info(f"Excel export muvaffaqiyatli: {filename}, {records_count} records, User: {user_id}")
        
        return {
            'success': True,
            'file_path': relative_path,
            'file_url': file_url,
            'filename': filename,
            'records_count': records_count,
        }
        
    except Exception as e:
        logger.error(f"Excel export xatolik: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': f'Export xatolik: {str(e)}'
        }


@shared_task(bind=True, name='finance.export_payments_to_excel')
def export_payments_to_excel(self, branch_id, filters=None, user_id=None):
    """
    To'lovlarni Excel fayliga export qilish.
    
    Args:
        branch_id: Filial ID
        filters: Filter parametrlari (dict)
        user_id: Export qiluvchi user ID
        
    Returns:
        dict: Export natijasi
    """
    try:
        from apps.school.finance.models import Payment
        from apps.branch.models import Branch
        from django.conf import settings
        
        # Branch ma'lumotini olish
        try:
            branch = Branch.objects.get(id=branch_id, deleted_at__isnull=True)
        except Branch.DoesNotExist:
            return {
                'success': False,
                'error': 'Filial topilmadi'
            }
        
        # To'lovlarni olish
        queryset = Payment.objects.filter(
            branch_id=branch_id,
            deleted_at__isnull=True
        ).select_related(
            'student_profile',
            'student_profile__user_branch__user',
            'subscription_plan',
            'discount',
            'transaction',
            'branch'
        ).order_by('-payment_date', '-created_at')
        
        # Filterlarni qo'llash
        if filters:
            if 'student_profile' in filters:
                queryset = queryset.filter(student_profile_id=filters['student_profile'])
            if 'date_from' in filters:
                queryset = queryset.filter(payment_date__gte=filters['date_from'])
            if 'date_to' in filters:
                queryset = queryset.filter(payment_date__lte=filters['date_to'])
            if 'period' in filters:
                queryset = queryset.filter(period=filters['period'])
        
        # Limitatsiya
        MAX_RECORDS = 50000
        records_count = queryset.count()
        
        if records_count > MAX_RECORDS:
            return {
                'success': False,
                'error': f'Juda ko\'p yozuv ({records_count}). Maksimum {MAX_RECORDS} yozuvgacha export qilish mumkin.'
            }
        
        if records_count == 0:
            return {
                'success': False,
                'error': 'Export qilish uchun ma\'lumot topilmadi'
            }
        
        # Excel fayl yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = "To'lovlar"
        
        # Headerlar
        headers = [
            '№',
            'Sana',
            'O\'quvchi',
            'Abonement',
            'Davr',
            'Asosiy summa',
            'Chegirma summasi',
            'Yakuniy summa',
            'Davr boshlanishi',
            'Davr tugashi',
            'Eslatma',
            'Yaratilgan',
        ]
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headerlarni yozish
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Ma'lumotlarni yozish
        for row_num, payment in enumerate(queryset, 2):
            # 1. Tartib raqam
            ws.cell(row=row_num, column=1, value=row_num - 1).border = border
            
            # 2. Sana
            date_value = payment.payment_date.strftime('%d.%m.%Y') if payment.payment_date else ''
            ws.cell(row=row_num, column=2, value=date_value).border = border
            
            # 3. O'quvchi
            student_name = ''
            if payment.student_profile and hasattr(payment.student_profile, 'user_branch'):
                user = payment.student_profile.user_branch.user
                student_name = f"{user.first_name} {user.last_name}".strip()
            ws.cell(row=row_num, column=3, value=student_name).border = border
            
            # 4. Abonement
            plan_name = payment.subscription_plan.name if payment.subscription_plan else ''
            ws.cell(row=row_num, column=4, value=plan_name).border = border
            
            # 5. Davr
            period_display = payment.get_period_display() if payment.period else ''
            ws.cell(row=row_num, column=5, value=period_display).border = border
            
            # 6. Asosiy summa
            base_cell = ws.cell(row=row_num, column=6, value=payment.base_amount)
            base_cell.number_format = '#,##0'
            base_cell.border = border
            
            # 7. Chegirma summasi
            discount_cell = ws.cell(row=row_num, column=7, value=payment.discount_amount)
            discount_cell.number_format = '#,##0'
            discount_cell.border = border
            
            # 8. Yakuniy summa
            final_cell = ws.cell(row=row_num, column=8, value=payment.final_amount)
            final_cell.number_format = '#,##0'
            final_cell.font = Font(bold=True)
            final_cell.border = border
            
            # 9. Davr boshlanishi
            period_start = payment.period_start.strftime('%d.%m.%Y') if payment.period_start else ''
            ws.cell(row=row_num, column=9, value=period_start).border = border
            
            # 10. Davr tugashi
            period_end = payment.period_end.strftime('%d.%m.%Y') if payment.period_end else ''
            ws.cell(row=row_num, column=10, value=period_end).border = border
            
            # 11. Eslatma
            notes = payment.notes or ''
            ws.cell(row=row_num, column=11, value=notes).border = border
            
            # 12. Yaratilgan
            created_date = payment.created_at.strftime('%d.%m.%Y %H:%M') if payment.created_at else ''
            ws.cell(row=row_num, column=12, value=created_date).border = border
        
        # Ustunlar kengligini moslash
        column_widths = {
            1: 8,   # №
            2: 12,  # Sana
            3: 30,  # O'quvchi
            4: 25,  # Abonement
            5: 12,  # Davr
            6: 15,  # Asosiy summa
            7: 18,  # Chegirma
            8: 15,  # Yakuniy summa
            9: 15,  # Davr boshi
            10: 15, # Davr oxiri
            11: 30, # Eslatma
            12: 18, # Yaratilgan
        }
        
        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Faylni saqlash
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"payments_{branch.name}_{timestamp}.xlsx"
        
        exports_dir = os.path.join(settings.MEDIA_ROOT, 'exports', 'finance')
        os.makedirs(exports_dir, exist_ok=True)
        
        file_path = os.path.join(exports_dir, filename)
        wb.save(file_path)
        
        relative_path = os.path.join('exports', 'finance', filename)
        file_url = f"{settings.MEDIA_URL}{relative_path}"
        
        logger.info(f"Payments Excel export muvaffaqiyatli: {filename}, {records_count} records, User: {user_id}")
        
        return {
            'success': True,
            'file_path': relative_path,
            'file_url': file_url,
            'filename': filename,
            'records_count': records_count,
        }
        
    except Exception as e:
        logger.error(f"Payments Excel export xatolik: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': f'Export xatolik: {str(e)}'
        }
